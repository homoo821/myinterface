# -*- coding:utf-8 -*-
#Author: homoo
#Data: 2020/2/24
#File: test_request.py

import unittest
import openpyxl
from ddt import ddt,data
from tools.read_excel import ReadExcel
from api_test.http_request import HttpRequest
from tools.file_path import *
from tools.get_data import GetData
from tools.user_log import *

test_data = ReadExcel().read_excel()

@ddt
class TestHttp(unittest.TestCase):

    def setUp(self):
        # print('==========开始测试==========')
        UserLog().info('==========开始测试==========')

    @data(*test_data)
    def test_http(self, test_data):

        # print('/*开始->ID:{0},标题:{1}*/'.format(test_data['ID'],test_data['title']))
        UserLog().info('/*开始->ID:{0},标题:{1}*/'.format(test_data['ID'],test_data['title']))
        # if test_data['data'].find("${name}") != -1:
        #     wb = openpyxl.load_workbook(test_data_path)
        #     sheet=wb['init']
        #     name = str(sheet.cell(2,1).value)
        #     test_data['data']=test_data['data'].replace('${name}',name)
        #     sheet.cell(2,1).value += 1
        #     wb.save(test_data_path)
        # else:
        #     test_data['data']=test_data['data']

        r = HttpRequest().http_request(test_data['url'], eval(test_data["header"]), eval(test_data['data']), test_data['method'], getattr(GetData,'Cookie'))
        # print('/*结束->ID:{0},标题:{1}*/'.format(test_data['ID'],test_data['title']))
        UserLog().info('/*结束->ID:{0},标题:{1}*/'.format(test_data['ID'],test_data['title']))
        if r.cookies:
            setattr(GetData, 'Cookie', r.cookies)
        ReadExcel().writ_back(test_data_path, test_data['sheet_name'], int(test_data['ID'])+1, 'res', str(r.json()))
        # try:
        #     if r.json()['retcode']==test_data['exp']:
        #         ReadExcel().writ_back(test_data_path, 'addteacher', int(test_data['ID'])+1, 'testresult', 'Pass')
        #     else:
        #         ReadExcel().writ_back(test_data_path, 'addteacher', int(test_data['ID'])+1, 'testresult', 'Fail')
        # except Exception as e:
        #     print(e)
        #     raise e
        try:
            self.assertEqual(r.json()['retcode'], test_data['exp'])
            testresult = 'Pass'
        except Exception as e:
            testresult = 'Fail'
            print(e)
            raise e
        finally:
            ReadExcel().writ_back(test_data_path, test_data['sheet_name'], int(test_data['ID']) + 1, 'testresult', testresult)

    def tearDown(self):
        # print('==========测试结束==========')
        UserLog().info('==========测试结束==========')
if __name__ == '__main__':
    unittest.main()

