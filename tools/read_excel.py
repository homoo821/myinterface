# -*- coding:utf-8 -*-
# Author: homoo
# Data: 2020/2/24
# File: read_excel.py

import openpyxl
from tools.file_path import *
from tools.read_config import ReadConfig


class ReadExcel():
    """读取Excel文件"""

    def read_excel(self):
        mode = eval(ReadConfig().read_config(config_path, 'MODE', 'mode'))
        wb = openpyxl.load_workbook(test_data_path)

        test_data=[]
        for key in mode:
            sheet = wb[key]
            if mode[key] == 'all':
                for row in range(2, sheet.max_row + 1):
                    each_data = {}
                    for col in range(1, sheet.max_column + 1):
                        each_data[sheet.cell(1, col).value] = sheet.cell(row, col).value
                    if each_data['data'].find("${name}") != -1:
                        name = str(wb['init'].cell(2, 1).value)
                        each_data['data'] = each_data['data'].replace('${name}', name)
                        wb['init'].cell(2, 1).value += 1
                        wb.save(test_data_path)
                    else:
                        each_data['data'] = each_data['data']
                    each_data['sheet_name']=key
                    test_data.append(each_data)

            else:
                for case_id in mode[key]:
                    each_data = {}
                    for col in range(1, sheet.max_column + 1):
                        each_data[sheet.cell(1, col).value] = sheet.cell(case_id+1, col).value
                    each_data['sheet_name'] = key
                    test_data.append(each_data)

        return test_data

    def writ_back(self, file_path, sheet_name, row, col_name, value):
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            print('文件打开异常:{}'.format(e))
            raise e
        sheet = wb[sheet_name]
        for col in range(1, sheet.max_column + 1):
            if col_name == sheet.cell(1, col).value:
                sheet.cell(row, col).value = value
            else:
                continue
        wb.save(file_path)


if __name__ == '__main__':
    r = ReadExcel().read_excel()
    print(r)
    print(len(r))
    # r2 = ReadExcel().writ_back("../data/test_data.xlsx","addteacher",4,'res','fasfa')